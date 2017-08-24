#!/usr/bin/env python
# -*- coding: utf-8 ; mode: python -*-
#
# Copyright (C) 2017 DA Symposium 2017
# Copyright (c) 2014,2015,2016 dasadc, Fujitsu
# All rights reserved.
#
# Redistribution and use in source and binary forms, with or without
# modification, are permitted provided that the following conditions
# are met:
# 1. Redistributions of source code must retain the above copyright
#    notice, this list of conditions and the following disclaimer.
# 2. Redistributions in binary form must reproduce the above copyright
#    notice, this list of conditions and the following disclaimer in the
#    documentation and/or other materials provided with the distribution.
#
# THIS SOFTWARE IS PROVIDED BY THE AUTHOR AND CONTRIBUTORS ``AS IS'' AND
# ANY EXPRESS OR IMPLIED WARRANTIES, INCLUDING, BUT NOT LIMITED TO, THE
# IMPLIED WARRANTIES OF MERCHANTABILITY AND FITNESS FOR A PARTICULAR PURPOSE
# ARE DISCLAIMED.  IN NO EVENT SHALL THE AUTHOR OR CONTRIBUTORS BE LIABLE
# FOR ANY DIRECT, INDIRECT, INCIDENTAL, SPECIAL, EXEMPLARY, OR CONSEQUENTIAL
# DAMAGES (INCLUDING, BUT NOT LIMITED TO, PROCUREMENT OF SUBSTITUTE GOODS
# OR SERVICES; LOSS OF USE, DATA, OR PROFITS; OR BUSINESS INTERRUPTION)
# HOWEVER CAUSED AND ON ANY THEORY OF LIABILITY, WHETHER IN CONTRACT, STRICT
# LIABILITY, OR TORT (INCLUDING NEGLIGENCE OR OTHERWISE) ARISING IN ANY WAY
# OUT OF THE USE OF THIS SOFTWARE, EVEN IF ADVISED OF THE POSSIBILITY OF
# SUCH DAMAGE.
#


"""
Excelファイル(.xlsx)で作成したパズルの問題データを、
ソルバー用のテキストデータ形式へ、変換して、出力する。
アルゴリズムデザインコンテスト(ADC)形式も出力する。
複数のワークシートが存在してもよい。
### 制限事項
旧Excelファイル(.xls)はサポートしない。

usage: ./nlconv.py [-h] INPUTFILE.xlsx ...

"""

from __future__ import print_function
import sys
import os
# ./lib/ をsys.pathの末尾に追加
sys.path.append(os.path.abspath(os.path.join(os.path.dirname(__file__), './lib')))
import re
from openpyxl import load_workbook
# python3でlong型
if sys.version_info > (3,):
    long = int
#import json


rule = 2017 # ADCルール（デフォルト値）


def get_ban_size(ban):
    "banの、縦、横サイズを求める"
    xmax = len(ban)
    ymax = len(ban[0])
    # python版では、すべて同じ大きさのはずなので、ymaxはこれでよいはず
    return (xmax, ymax)


def print_ban(ban, fp, csv=False):
    "盤を出力する"
    xmax, ymax = get_ban_size(ban)
    spacer = ' '
    if csv:
        spacer = ','
    fp.write('%d\n' % ymax)
    fp.write('%d\n' % xmax)
    for y in range(0, ymax):
        for x in range(0, xmax):
            if ban[x][y] in (None, ' ', '-', 0):
                fp.write(' -')
            else:
                fp.write('%2d' % int(ban[x][y]))
            fp.write(spacer)
        fp.write('\n')

def print_ban2016(data, fp, csv=False):
    "盤を出力する (2016年版)"
    nlayers = len(data)
    xmax0 = None
    ymax0 = None
    spacer = ' '
    if csv:
        spacer = ','
    for i in range(1, nlayers+1):
        ban  = data[i]['ban']
        vias = data[i]['via']
        xmax, ymax = get_ban_size(ban)
        clear_number(ban) # 2回以上実行しても問題ない
        if i == 1:
            fp.write('%d\n%d\n' % (ymax, xmax))
            xmax0 = xmax
            ymax0 = ymax
        else:
            if (xmax0 != xmax) or (ymax0 != ymax):
                raise Exception('ERROR: ban size must be same: %dX%d != %dX%d' % (xmax0, ymax0, xmax, ymax))
        fp.write('LAYER %d\n' % i)
        # 盤面データを出力する
        for y in range(0, ymax):
            for x in range(0, xmax):
                if ban[x][y] in (None, ' ', '-', 0):
                    fp.write(' -')
                else:
                    via = find_via(vias, x, y, i)
                    #print('via=', via)
                    if via == '': # 数字マスのとき
                        fp.write('%02d' % ban[x][y])
                    else:
                        fp.write('%2s' % via)
                fp.write(spacer)
            fp.write('\n')


def print_ban2017(data, fp, csv=False):
    "盤を出力する (2017年版)"
    nlayers = len(data)
    xmax0 = None
    ymax0 = None
    spacer = ' '
    if csv:
        spacer = ','
    clear_number_3D(data)
    for i in range(1, nlayers+1):
        ban  = data[i]['ban']
        vias = data[i]['via']
        xmax, ymax = get_ban_size(ban)
        if i == 1:
            fp.write('%d\n%d\n' % (ymax, xmax))
            xmax0 = xmax
            ymax0 = ymax
        else:
            if (xmax0 != xmax) or (ymax0 != ymax):
                raise Exception('ERROR: ban size must be same: %dX%d != %dX%d' % (xmax0, ymax0, xmax, ymax))
        fp.write('LAYER %d\n' % i)
        # 盤面データを出力する
        for y in range(0, ymax):
            for x in range(0, xmax):
                if ban[x][y] in (None, ' ', '-', 0):
                    fp.write(' -')
                else:
                    fp.write('%02d' % ban[x][y])
                fp.write(spacer)
            fp.write('\n')


def print_ban_adc(ban, num, fp):
    "盤を出力する (ADC形式)"
    xmax, ymax = get_ban_size(ban)
    if num == 1:
        fp.write('SIZE %dX%d\n' % (xmax, ymax))
    else:
        fp.write('\n')
    for y in range(0, ymax):
        for x in range(0, xmax):
            if ban[x][y] in (None, ' ', '-'):
                ban[x][y] = 0
            fp.write('%02d' % int(ban[x][y]))
            if x < xmax-1:
                fp.write(',')
        fp.write('\n')


def print_ban_adc2016(data, fp):
    "盤を出力する (ADC2016形式, 回答データ)"
    nlayers = len(data)
    xmax0=None
    ymax0=None
    for i in range(1, nlayers+1):
        ban = data[i]['ban']
        via = data[i]['via']
        xmax, ymax = get_ban_size(ban)
        if i == 1:
            fp.write('SIZE %dX%dX%d\n' % (xmax, ymax, nlayers))
            xmax0 = xmax
            ymax0 = ymax
        else:
            if (xmax0 != xmax) or (ymax0 != ymax):
                raise Exception("ERROR: ban size must be same: %dX%d != %dX%d" % (xmax0, ymax0, xmax, ymax))
        fp.write('LAYER %d\n' % i)
        # 盤面データを出力する
        for y in range(0, ymax):
            for x in range(0, xmax):
                if ban[x][y] in (None, ' ', '-'):
                    ban[x][y] = 0
                fp.write('%02d' % ban[x][y])
                if x < xmax-1:
                    fp.write(',')
            fp.write('\n')


def find_via(vias, x, y, layer):
    "ビアの名前を検索して返す。"
    for via in sorted(vias.keys()):
        if ((vias[via]['x'] == x) and
            (vias[via]['y'] == y) and
            (vias[via]['z'] == layer)):
            return via
    return ''
        

def clear_number(ban):
    "リンク途中の余計な数字を消して、リンクの端点のみを残す"
    xmax, ymax = get_ban_size(ban)
    #print('clear_number: xmax=%d, ymax=%d' % (xmax, ymax))
    #print_ban(ban, sys.stdout)
    # まずは、消す候補を数え上げる
    clear = [[' ']*ymax for i in range(xmax)]
    for y in range(0, ymax):
        for x in range(0, xmax):
            num = ban[x][y]
            if num in ('-', ' '):
                continue
            # 四方に隣接するマスと、一致する個数が２以上なら、消してよい？
            # ---> 隣接するマスが１個だけ（＝端点）は消してはいけない。
            count = 0
            #上のマスと同じ値か？ …文字列比較で？？ 数値, " ", "-"
            if (0 <= y-1) and (ban[x][y-1] == num):
                count += 1
            #左のマスと同じ値か？
            if (0 <= x-1) and (ban[x-1][y] == num):
                count += 1
            #下のマスと同じ値か？ 
            if (y+1 < ymax) and (ban[x][y+1] == num):
                count += 1
            #右のマスと同じ値か？
            if (x+1 < xmax) and (ban[x+1][y] == num):
                count += 1
            if 2 <= count:
                clear[x][y] = 1 # 消す
            else:
                clear[x][y] = 0 # 消さない
            #print(x, y, count)
    #print('clear='); print_ban(clear, sys.stdout)
    #本当に消す
    for y in range(0, ymax):
        for x in range(0, xmax):
            if clear[x][y] == 1:
                ban[x][y] = ' '
    #print_ban(ban, sys.stdout)


def clear_number_3D(data):
    "リンク途中の余計な数字を消して、リンクの端点のみを残す (3次元版)"
    nlayers = len(data)
    ban1 = data[1]['ban']
    xmax, ymax = get_ban_size(ban1)
    # 隣接ノード（N,E,W,S,U,D）に、同じ数字のマスがいくつあるか数える
    same = [[[0]*ymax for x in range(xmax)] for z in range(nlayers+1)] # [z][x][y]
    for z in range(1, nlayers+1):
        ban = data[z]['ban']
        # U (上の層)
        if z+1 <=nlayers:
            banU = data[z+1]['ban']
        else:
            banU = None
        # D (下の層)
        if 1 <= z-1:
            banD = data[z-1]['ban']
        else:
            banD = None
        #print('LAYER %d' % z); print_ban(ban, sys.stdout)
        for x in range(0, xmax):
            for y in range(0, ymax):
                num = ban[x][y]
                if num in ('-', ' ', 0):
                    continue
                count = 0
                # N(上)のマスと同じ値か？ 
                if (0 <= y-1) and (ban[x][y-1] == num):
                    count += 1
                # W(左)のマスと同じ値か？
                if (0 <= x-1) and (ban[x-1][y] == num):
                    count += 1
                # S(下)のマスと同じ値か？ 
                if (y+1 < ymax) and (ban[x][y+1] == num):
                    count += 1
                # E(右)のマスと同じ値か？
                if (x+1 < xmax) and (ban[x+1][y] == num):
                    count += 1
                # U(上層)のマスと同じ値か？
                if banU and (banU[x][y] == num):
                    count += 1
                # DU(下層)のマスと同じ値か？
                if banD and (banD[x][y] == num):
                    count += 1
                same[z][x][y] = count
                #print('%s (%d,%d,%d) = %d' % (num, x,y,z, count))
    #for z in range(1, nlayers+1):
    #    print('layer=%d, same=' % z); print_ban(same[z], sys.stdout)
    #本当に消す
    for z in range(1, nlayers+1):
        ban = data[z]['ban']
        for x in range(0, xmax):
            for y in range(0, ymax):
                    if 3 <= same[z][x][y]: # 分岐がある
                        print('WARNING: branch (%d,%d,%d) #same=%d' % (x,y,z, same[z][x][y]))
                    if 2 <= same[z][x][y]:
                        ban[x][y] = ' '
        #print('LAYER %d' % z); print_ban(ban, sys.stdout)


def enumerate_numbers(ban, xmax, ymax):
    "番号を数え上げる"
    numbers = {}  # key=数字マスの数字, value=(数字マスの個数)
    pos0 = {}     # key=0番目の数字マスの数字, value=マスの座標"(x,y)"
    pos1 = {}     # key=1番目の数字マスの数字, value=マスの座標"(x,y)"
    maxNum = -1   # 数字マスの数字の、最大値
    #print_ban(ban, sys.stdout)
    for y in range(0, ymax):
        for x in range(0, xmax):
            #print(type(ban[x][y]), ban[x][y])
            if type(ban[x][y]) is int:
                num = ban[x][y]
                if num == 0: # ???
                    continue
                if maxNum <= num: maxNum = num
                if numbers.get(num) is None:
                    numbers[num] = 0
                    pos0[num] = '(%d,%d)' % (x,y)
                if numbers[num] == 1:
                    pos1[num] = '(%d,%d)' % (x,y)
                if numbers[num] >= 2:
                    raise Exception('ERROR: duplicated number %d (%d)' % (num, numbers[num]+1))
                numbers[num] += 1
    return {'numbers': numbers,
            'pos0': pos0,
            'pos1': pos1,
            'maxNum': maxNum}


def print_Q_adc(ban, fp):
    "ADC形式の出題ファイルを書き出す"
    xmax, ymax = get_ban_size(ban)
    fp.write('SIZE %dX%d\n' % (xmax, ymax))

    # 番号を数え上げる
    en = enumerate_numbers(ban, xmax, ymax)
    numbers = en['numbers']
    pos0    = en['pos0']
    pos1    = en['pos1']
    maxNum  = en['maxNum']

    # 出力する
    fp.write('LINE_NUM %d\n' % maxNum)
    for i in range(1, maxNum+1): # 線の番号は1から始まる
        if numbers.get(i) is None:
            fp.write('ERROR: cannot find number %d\n' % i)
            raise Exception('ERROR: cannot find number %d' % i)
        else:
            if pos1.get(i) is None:
                raise Exception("ERROR: cannot find second number %d" % i)
            fp.write('LINE#%d %s-%s\n' % (i, pos0[i], pos1[i]))


def print_Q_adc2016(data, fp):
    "ADC形式の出題ファイルを書き出す (2016年版)"
    nlayers = len(data)
    xmax0 = None
    ymax0 = None
    maxNum = 0
    enum_num = {}
    for i in range(1, nlayers+1):
        ban  = data[i]['ban']
        xmax, ymax = get_ban_size(ban)
        clear_number(ban) # 2回以上実行しても問題ない
        if i == 1:
            xmax0 = xmax
            ymax0 = ymax
        else:
            # 盤面サイズが一致していることを確認
            if (xmax0 != xmax) or (ymax0 != ymax):
                raise Exception("ERROR: ban size must be same: %dX%d != %dX%d" % (xmax0, ymax0, xmax, ymax))
        # 番号を数え上げる
        en = enumerate_numbers(ban, xmax, ymax)
        if maxNum < en['maxNum']: maxNum = en['maxNum']
        enum_num[i] = en
    numbers0 = {}
    lines = {}
    viaList = {}
    fp.write('SIZE %dX%dX%d\n' % (xmax0, ymax0, nlayers))
    fp.write('LINE_NUM %d\n' % maxNum)
    for layer in range(1, nlayers+1):
        en = enum_num[layer]
        #print(en)
        numbers = en['numbers']
        pos0    = en['pos0']
        pos1    = en['pos1']
        for line in range(1, maxNum+1):
            if numbers.get(line) is not None:
                if numbers0.get(line) is not None:
                    #print('1: numbers0[%d] = %d' % (line, numbers0[line]))
                    numbers0[line] += numbers[line] # 番号マスの出現回数を加算
                    #print('2: numbers0[%d] = %d' % (line, numbers0[line]))
                else:
                    numbers0[line] = numbers[line]
                    #print('0: numbers0[%d] = %d' % (line, numbers0[line]))
            if pos0.get(line) is not None:
                if pos1.get(line) is None:
                    through_via = is_through_via(data, line, layer)
                    if through_via is not None:
                        viaName = through_via['name']
                        if viaList.get(viaName) is None: viaList[viaName] = ''
                        viaList[viaName] += ' (%d,%d,%d)' % (through_via['x'], through_via['y'], through_via['z'])
                    else:
                        raise Exception('ERROR: cannot find second number of LINE %d' % line)
                    continue # いずれにしても、処理続行
                x0 = None; y0 = None; x1 = None; y1 = None
                m = re.match(r'\((\d+),(\d+)\)', pos0[line])
                if m is None:
                    raise Exception('BUG: format error: LINE %d pos0=%s' % (line, pos0[line]))
                else:
                    x0 = int(m.group(1))
                    y0 = int(m.group(2))
                m = re.match(r'\((\d+),(\d+)\)', pos1[line])
                if m is None:
                    raise Exception('BUG: format error: LINE %d pos1=%s' % (line, pos1[line]))
                else:
                    x1 = int(m.group(1))
                    y1 = int(m.group(2))

                via0 = find_via(data[layer]['via'], x0, y0, layer)
                via1 = find_via(data[layer]['via'], x1, y1, layer)
                if via0 == '':
                    if lines.get(line) is None: lines[line] = {}
                    if lines[line].get('pos0') is not None:
                        if lines[line].get('pos1') is not None:
                            # 複数のlayerで、個別に、同じ番号の線が引かれているとき
                            raise Exception('ERROR: too many number %d: %s and (%d,%d,%d)' % (line, lines[line]['pos0'], x0, y0, layer))
                        else:
                            # pos0がすでに決まっていたので、pos1に登録
                            lines[line]['pos1'] = '(%d,%d,%d)' % (x0, y0, layer)
                    else:
                        lines[line]['pos0'] = '(%d,%d,%d)' % (x0, y0, layer)
                else:
                    #print('VIA0 (%d,%d,%d) for LINE#%d' % (x0, y0, layer, line))
                    if viaList.get(via0) is None: viaList[via0] = ''
                    viaList[via0] += ' (%d,%d,%d)' % (x0, y0, layer)
                if via1 == '':
                    if lines.get(line) is None: lines[line] = {}
                    if lines[line].get('pos1') is not None:
                        if lines[line].get('pos0') is not None:
                            # 複数のlayerで、個別に、同じ番号の線が引かれているとき
                            raise Exception('ERROR: too many number %d: %s and (%d,%d,%d)' % (line, lines[line]['pos1'], x1, y1, layer))
                        else:
                            # pos1がすでに決まっていたので、pos0に登録
                            lines[line]['pos0'] = '(%d,%d,%d)' % (x1, y1, layer)
                    else:
                        lines[line]['pos1'] = '(%d,%d,%d)' % (x1, y1, layer)
                else:
                    #print('VIA1 (%d,%d,%d) for LINE#%d' % (x1, y1, layer, line))
                    if viaList.get(via1) is None: viaList[via1] = ''
                    viaList[via1] += ' (%d,%d,%d)' % (x1, y1, layer)
                #print('LINE %d (%d,%d,%d)-(%d,%d,%d) ' % (line, x0, y0, layer, x1, y1, layer))
    for line in range(1, maxNum+1):
        if numbers0.get(line) is not None:
            #print('numbers0[%d] = %d' % (line, numbers0[line]))
            fp.write('LINE#%d %s %s\n' % (line, lines[line]['pos0'], lines[line]['pos1']))
        else:
            fp.write('ERROR: cannot find number %d\n' % line)
            raise Exception('ERROR: cannot find number %d' % line)
    for viaName in sorted(viaList.keys()):
        fp.write('VIA#%s%s\n' % (viaName, viaList[viaName]))


def print_Q_adc2017(data, fp):
    "ADC形式の出題ファイルを書き出す (2017年版)"
    #print(json.dumps(data, sort_keys=True, indent=4)); sys.exit(1)
    nlayers = len(data)
    xmax0 = None
    ymax0 = None
    maxNum = 0
    enum_num = {} # key=層の番号, value=enumerate_numbersの結果
    clear_number_3D(data)
    for i in range(1, nlayers+1):
        ban  = data[i]['ban']
        xmax, ymax = get_ban_size(ban)
        if i == 1:
            xmax0 = xmax
            ymax0 = ymax
        else:
            # 盤面サイズが一致していることを確認
            if (xmax0 != xmax) or (ymax0 != ymax):
                raise Exception("ERROR: ban size must be same: %dX%d != %dX%d" % (xmax0, ymax0, xmax, ymax))
        # 番号を数え上げる
        en = enumerate_numbers(ban, xmax, ymax)
        if maxNum < en['maxNum']: maxNum = en['maxNum']
        enum_num[i] = en
    numbers0 = {} # key=数字マスの数字, value=番号マスの出現回数
    lines = {}    # key=数字マスの数字
    viaList = {}  # key=viaの名前

    def register_lines(x, y, layer):
        "lines[番号]に、座標を登録する"
        if lines.get(line) is None: lines[line] = {} # 初期化
        if lines[line].get('pos0') is None:
            lines[line]['pos0'] = '(%d,%d,%d)' % (x, y, layer)
        elif lines[line].get('pos1') is None:
            # pos0がすでに決まっていたので、pos1に登録
            lines[line]['pos1'] = '(%d,%d,%d)' % (x, y, layer)
        else:
            # 複数のlayerで、個別に、同じ番号の線が引かれているとき ?????
            raise Exception('ERROR: too many number %d: %s %s (%d,%d,%d)' % (line, lines[line]['pos0'], lines[line]['pos1'], x, y, layer))


    fp.write('SIZE %dX%dX%d\n' % (xmax0, ymax0, nlayers))
    fp.write('LINE_NUM %d\n' % maxNum)
    for layer in range(1, nlayers+1):    # レイヤーについて...
        en = enum_num[layer]
        #print('layer=%d' % layer, 'en=', en)
        numbers = en['numbers']
        pos0    = en['pos0']
        pos1    = en['pos1']
        for line in range(1, maxNum+1):  # 数字マスの数字について...
            x0 = None; y0 = None; x1 = None; y1 = None
            # 数字マスの出現回数を数える
            if numbers.get(line) is not None:
                if numbers0.get(line) is None:
                    numbers0[line] = numbers[line]
                else:
                    numbers0[line] += numbers[line]
            # 数字マスの座標を決定する
            if pos0.get(line) is not None:
                m = re.match(r'\((\d+),(\d+)\)', pos0[line]) # なんだかバカバカしい。あとで直す
                if m is None:
                    raise Exception('BUG: format error: LINE %d pos0=%s' % (line, pos0[line]))
                x0 = int(m.group(1))
                y0 = int(m.group(2))
                #via0 = find_via(data[layer]['via'], x0, y0, layer)
                #print('pos0: [%d] (%d,%d,%d):"%s"' % (line, x0, y0, layer, via0))
                register_lines(x0, y0, layer)
            if pos1.get(line) is not None:
                m = re.match(r'\((\d+),(\d+)\)', pos1[line])
                if m is None:
                    raise Exception('BUG: format error: LINE %d pos1=%s' % (line, pos1[line]))
                x1 = int(m.group(1))
                y1 = int(m.group(2))
                #via1 = find_via(data[layer]['via'], x1, y1, layer)
                #print('pos1: [%d] (%d,%d,%d):"%s"' % (line, x1, y1, layer, via1))
                register_lines(x1, y1, layer)

    for line in range(1, maxNum+1):
        if numbers0.get(line) is not None:
            #print('numbers0[%d] = %d' % (line, numbers0[line]))
            fp.write('LINE#%d %s %s\n' % (line, lines[line]['pos0'], lines[line]['pos1']))
        else:
            fp.write('ERROR: cannot find number %d\n' % line)
            raise Exception('ERROR: cannot find number %d' % line)


def dump_vias(vias):
    "ビアの情報をダンプ出力する。"
    for via in sorted(vias.keys()):
        print('%s (%d,%d,%d)=%d' % (via, vias[via]['x'], vias[via]['y'], vias[via]['z'], vias[via]['val']))


def check_via(data):
    """
    viaの(X,Y)座標が一致するかチェックする。
    中間のLAYERでviaが置かれているかチェックする。
    viaだけがあって、数字マスと接続していない場合は…？  別のチェック条件で検出されるはず
    """
    nlayers = len(data)
    found_name = {}
    found_xyz = {}
    err = 0
    for i in range(1, nlayers+1):
        vias = data[i]['via']
        for via in sorted(vias.keys()):
            x = vias[via]['x']
            y = vias[via]['y']
            z = vias[via]['z']
            xy  = 'x=%d,y=%d' % (x, y)
            xyz = 'x=%d,y=%d,z=%d' % (x, y, z)
            # (X,Y)座標が一致すること
            if found_name.get(via) is not None:
                if found_name[via] == xy:
                    # (X,Y)が一致したのでOK
                    pass
                else:
                    print('ERROR: VIA#%s X,Y mismatch: %s and %s' % (via, found_name[via], xy))
                    err += 1
            else:
                # 初登場
                found_name[via] = xy
            # 中間のLAYERでviaが置かれていること
            if found_xyz.get(xyz) is not None:
                if found_xyz[xyz] == i-1: # layerが連続している
                    # ok
                    pass
                else:
                    print('ERROR: VIA#%s(%d,%d,%d) is missing' % (via, x, y, i-1))
                    err += 1
            else:
                # 初登場
                found_xyz[xyz] = i
    return err


def via_name(i):
    """
    数値から、viaの名前を作る。a b c ... z aa ab ac ...
    @param i 正の整数 0, 1, 2, ...
    """
    tbl = 'abcdefghijklmnopqrstuvwxyz'
    N = len(tbl)
    if i < N:
        return tbl[i]
    else:
        j = i % N
        return via_name(int(i/N)-1) + tbl[j]

def detect_vias(data):
    "(ADC2017ルール用) viaを探す"
    nlayers = len(data)
    ban1 = data[1]['ban']
    xmax, ymax = get_ban_size(ban1)
    enum_via = {}
    for x in range(0, xmax):
        for y in range(0, ymax):
            for i in range(2, nlayers+1):
                ban = data[i]['ban']
                num0 = data[i]['ban'][x][y]   # 今見ているレイヤー
                num1 = data[i-1]['ban'][x][y] # 1つ下のレイヤー
                if (num0 in ('-', ' ')) or (num1 in ('-', ' ')):
                    continue
                if num0 == num1:
                    key = '(%d,%d)' % (x,y)
                    if key not in enum_via:
                        enum_via[key] = {}
                    enum_via[key][i-1] = ((x,y,i-1), num0)
                    enum_via[key][i]   = ((x,y,i),   num0)
    #print('enum_via=', enum_via)
    vias = {}
    nvia = 0
    for key in sorted(enum_via.keys()):
        name = via_name(nvia)
        nvia += 1
        desc = 'VIA#%s ' % name
        for layer in sorted(enum_via[key].keys()):
            xyz, num = enum_via[key][layer]
            desc += str(xyz)
            data[layer]['via'][name] = {'x': xyz[0], 'y': xyz[1], 'z': xyz[2],
                                        'val': num, 'name': name}
        desc += ' //' + str(num)
        #print(desc)


def is_through_via(data, line, layer):
    "LAYERを通過するviaか？ （厳密なチェックではない）"
    nlayers = len(data)
    if layer == 1 or layer == nlayers: # 最下位層 or 最上位層
        return None # 違う
    vias = data[layer]['via']
    for via in sorted(vias.keys()):
        if vias[via]['val'] == line:
            return vias['via'] # 通過するviaである（可能性がある。もしもそこから線がのびてたら、ADC2016のルール違反）
    return None # 違う


def get_ban_data(worksheet, nrows, ncols, layer):
    """
    盤のデータを取り出す。
    ワークシートは、1行、1列(A列)から始まる。左上の座標が(row=1,colum=1)

    1行目:           属性情報: 000 行
    2行目:           属性情報: 000 列
    3行目:           列番号: 0 1 2 3 ...
    4行目:           盤のデータは、ここから始まる
    (4+nrows-1)行目: 盤のデータは、ここまで

    1列目:           行番号: 0 1 2 3 ...
    2列目:           盤のデータは、ここから始まる
    (2+ncols-1)列目: 盤のデータは、ここまで

    """
    xmax = 2 + ncols - 1
    ymax = 4 + nrows - 1
    ban2 = [[' ']*nrows for i in range(ncols)]
    vias = {}
    for y in range(4, ymax+1):
        for x in range(2, xmax+1):
            val = worksheet.cell(row=y, column=x).value
            #print(type(val), val)
            if val is None:
                ban2[x-2][y-4] = '-'
            elif type(val) is long:
                val = int(val)
                ban2[x-2][y-4] = val
            elif type(val) is unicode:
                val = str(val).replace(' ', '') # 余分なスペースを削除
                tmp = val.split('=')
                if len(tmp) == 2:
                    val = int(tmp[1])
                    vias[tmp[0]] = {'x': x-2,
                                    'y': y-4,
                                    'z': layer,
                                    'val': val,
                                    'name': tmp[0]}
                ban2[x-2][y-4] = val
            else:
                print('??? (%d,%d) %s' % (x,y, type(val)), end=''); print(val)
    return ban2, vias

def is_completed(multiData, worksheetbase, nlayers, diag=False):
    "すべてのレイヤーのデータがそろっている"
    complete = 0
    for i in range(1, nlayers+1):
        if i in multiData[worksheetbase]:
            complete += 1
    #print('is_completed: %s, nlayers=%d compete=%d' % (worksheetbase, nlayers, complete))
    if diag:
        print('%s [%d/%d]' % (worksheetbase, complete, nlayers))
    return (complete == nlayers)

def write_file(multiData, basefile, worksheetbase, dataA, layer_format):
    tmp = multiData[worksheetbase]
    nlayers = len(tmp)
    #print('nlayers=%d' % nlayers)
    ban = multiData[worksheetbase][1]['ban']
    via = multiData[worksheetbase][1]['via']
    data = multiData[worksheetbase]
    if dataA != 0:
        # 解答ファイルを作る(ソルバで解けなかったとき用)
        sfilename = '%s_%s_adc_sol.txt' % (basefile, worksheetbase)
        print(sfilename)
        with open(sfilename, 'w') as fp0:
            if nlayers == 1:
                print_ban_adc(ban, 1, fp0)
            else:
                print_ban_adc2016(data, fp0)
    # 線を消して、端点のみを残す
    if nlayers == 1:
        clear_number(ban)

    for ext,csv in [('txt', False), ('csv', True)]:
        ofilename = '%s_%s.%s' % (basefile, worksheetbase, ext)
        print(ofilename)
        with open(ofilename, 'w') as fp:
            if nlayers == 1:
                print_ban(ban, fp, csv=csv)
            elif rule == 2017:
                print_ban2017(data, fp, csv=csv)
            else:
                print_ban2016(data, fp, csv=csv)

    ofilename_adc = '%s_%s_adc.txt' % (basefile, worksheetbase)
    print(ofilename_adc)
    with open(ofilename_adc, 'w') as fp2:
        if layer_format == 0:
            print_Q_adc(ban, fp2)
        elif rule == 2017:
            print_Q_adc2017(data, fp2)
        else:
            print_Q_adc2016(data, fp2)


def proc1_sheet(basefile, worksheet, multiData):
    "ワークシート1枚を処理する"
    worksheetname = worksheet.title
    worksheetbase = worksheetname
    dataA = 0 # Aファイル（回答データ）を生成するか？ セルC1にAと書く
    nrows   = worksheet['A1'].value # 行数  "0000","行"  セルA1
    ncols   = worksheet['A2'].value # 列数  "0000","列"  セルA2
    layer   = worksheet['D1'].value # 層の番号           セルD1
    nlayers = worksheet['F1'].value # 層数               セルF1
    layer_format = 0                # 1 = 層が定義されている(2016年ルール)
    if (nrows is None) or (ncols is None):
        return None
    if (layer is not None) and (nlayers is not None):
        tmp = worksheetname.split('.')
        if len(tmp) != 2:
            raise Exception('ERROR: check worksheet name format: worksheet name=%s' % worksheetname)
        worksheetbase = tmp[0]
        if int(tmp[1]) != layer:
            raise Exception('ERROR: check layer numer: worksheet name=%s, layer=%d' % (worksheetname, layer))
        if layer <= 0 or nlayers < layer:
            raise Exception('ERROR: layer %d is out of range' % layer)
        layer_format = 1 # 2016年ルール
    elif (layer is not None) and (nlayers is None):
        raise Exception('ERROR: missing number of layer. check cell F1')
    elif (layer is None) and (nlayers is not None):
        raise Exception('ERROR: missing layer. check cell D1')
    else:
        layer = 1
        nlayers = 1
        layer_format = 0 # 2015年以前のルール
    if nrows <= 0 or ncols <= 0:
        print('SKIP: sheet %s' % worksheetname)
        return None
    #print('nrows=', nrows, 'ncols=', ncols)
    if worksheet['C1'].value == 'A':
        dataA = 1
    ban2, vias = get_ban_data(worksheet, nrows, ncols, layer)
    #print_ban(ban2, sys.stdout)
    #dump_vias(vias)
    if worksheetbase not in multiData:
        multiData[worksheetbase] = {}
    multiData[worksheetbase][layer] = {'ban': ban2, 'via': vias}
    complete = is_completed(multiData, worksheetbase, nlayers, diag=True)
    if complete:
        if rule == 2017:
            detect_vias(multiData[worksheetbase])
        check_via(multiData[worksheetbase])
        write_file(multiData, basefile, worksheetbase, dataA, layer_format)

def proc1(input_file):
    wb = load_workbook(filename=input_file)
    multiData = {}
    basefile = os.path.splitext(os.path.basename(input_file))[0]
    for worksheet in wb.worksheets:
        proc1_sheet(basefile, worksheet, multiData)


def main():
    global rule
    import argparse
    parser = argparse.ArgumentParser(description='Number Link data converter')
    parser.add_argument('-r', '--rule', choices=[2016, 2017], default=rule, type=int, help='set ADC rule (default: %(default)s)')
    parser.add_argument('file', nargs='*', help='input file (.xlsx)')
    args = parser.parse_args()
    rule = args.rule
    if args.file:
        for i in args.file:
            input_file = i
            print('input_file = %s' % input_file)
            proc1(input_file)
    else:
        parser.print_usage()

if __name__ == "__main__":
    main()
